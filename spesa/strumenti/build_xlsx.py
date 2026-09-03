# -*- coding: utf-8 -*-
import json, os, glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from dati import PRODOTTI, VOLANTINI, D

ARIAL  = 'Arial'
H_FILL = PatternFill('solid', fgColor='C00000')
H_FONT = Font(name=ARIAL, bold=True, color='FFFFFF', size=11)
TITLE  = Font(name=ARIAL, bold=True, size=14)
SUB    = Font(name=ARIAL, bold=True, size=12)
BODY   = Font(name=ARIAL, size=11)
NOTE   = Font(name=ARIAL, size=10, italic=True, color='666666')
THIN   = Border(bottom=Side('thin', color='DDDDDD'))
GIALLO = PatternFill('solid', fgColor='FFF2CC')

PDF   = {c: f for c, _, _, f in VOLANTINI}
PERIODO = {c: p for c, _, p, _ in VOLANTINI}

wb = Workbook()

# ---------------- Come si usa ----------------
ws = wb.active; ws.title = 'Come si usa'
for i,(t,f) in enumerate([
 ("Offerte dei supermercati vicino a corso Siracusa, Torino", TITLE),
 ("Aggiornato al 2 settembre 2026", NOTE), ("", BODY),
 ("Cosa trovi in questo file", SUB), ("", BODY),
 ("Foglio «I tuoi prodotti» — carne di bue, tonno e salmone in offerta, col prezzo al chilo", BODY),
 ("già calcolato. Ordinati dal più conveniente al più caro.", BODY), ("", BODY),
 ("Foglio «Indice pagine» — tutte le pagine degli 8 volantini, con le parole che il computer", BODY),
 ("è riuscito a leggerci dentro. Serve per cercare un prodotto qualsiasi: metti il filtro", BODY),
 ("sull'ultima colonna, scrivi per esempio caffè, e ti dice insegna e numero di pagina.", BODY),
 ("Poi apri quel PDF a quella pagina e leggi il prezzo.", BODY), ("", BODY),
 ("Due avvertenze", SUB), ("", BODY),
 ("1) Le righe gialle vengono da riassunti online e possono essere sbagliate: di errori così", BODY),
 ("   ne ho già trovati tre. Tutte le altre le ho lette dalla pagina del volantino.", BODY), ("", BODY),
 ("2) Le parole dell'indice le ha lette il computer dalle immagini: sono spesso storpiate e", BODY),
 ("   i prezzi non ci sono quasi mai. L'indice trova la pagina, il prezzo si legge sul PDF.", BODY), ("", BODY),
 ("Tessere del supermercato", SUB), ("", BODY),
 ("Alcuni prezzi valgono solo con la tessera: Ipercoop (soci Coop), Lidl Plus, Bennet Club.", BODY),
 ("Dove è così, sta scritto nelle note.", BODY), ("", BODY),
 ("Chi manca", SUB), ("", BODY),
 ("Mercatò: il loro sito non pubblica il volantino in un formato che si riesca a scaricare.", BODY),
 ("Carrefour Market: tolto su tua richiesta.", BODY),
], start=1):
    ws.cell(row=i, column=1, value=t).font = f
ws.column_dimensions['A'].width = 100

# ---------------- I tuoi prodotti ----------------
ws = wb.create_sheet('I tuoi prodotti')
COLS = ['Categoria','Insegna','Volantino valido','Reparto','Prodotto','Formato',
        'Quantità (kg)','Prezzo (€)','€ al kg','File PDF','Pag.','Fonte','Note']
for j,h in enumerate(COLS, start=1):
    c = ws.cell(row=1, column=j, value=h); c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(vertical='center', wrap_text=True)
for i,p in enumerate(PRODOTTI, start=2):
    cat,ins,chiave,rep,pro,fmt,qta,pre,pag,fon,note = p
    for j,v in enumerate([cat,ins,PERIODO[chiave],rep,pro,fmt,qta,pre,None,PDF[chiave],pag,fon,note], start=1):
        c = ws.cell(row=i, column=j, value=v); c.font = BODY; c.border = THIN
        c.alignment = Alignment(vertical='top', wrap_text=(j in (5,10,13)))
    ws.cell(row=i, column=9).value = f'=IFERROR(H{i}/G{i},"")'
    ws.cell(row=i, column=7).number_format = '0.000'
    ws.cell(row=i, column=8).number_format = '#,##0.00 "€"'
    c9 = ws.cell(row=i, column=9)
    c9.number_format = '#,##0.00 "€"'; c9.font = Font(name=ARIAL, size=11, bold=True)
    if fon == D:
        for j in range(1,14): ws.cell(row=i, column=j).fill = GIALLO
for col,w in zip('ABCDEFGHIJKLM',[13,15,34,13,48,24,12,11,11,34,7,20,52]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:M{len(PRODOTTI)+1}'
ws.cell(row=len(PRODOTTI)+3, column=1,
        value="«€ al kg» è calcolato: prezzo diviso quantità. Le righe gialle vanno controllate sul PDF.").font = NOTE

# ---------------- Indice pagine ----------------
ws = wb.create_sheet('Indice pagine')
idx = json.load(open('indice.json', encoding='utf-8'))
validi = {(os.path.basename(os.path.dirname(f)), int(os.path.basename(f)[:-4])) for f in glob.glob('pg/*/*.jpg')}
idx = [r for r in idx if (r['chiave'], r['pagina']) in validi]
idx.sort(key=lambda r: (r['insegna'], r['validita'], r['pagina']))
for j,h in enumerate(['Insegna','Volantino valido','File PDF','Pagina','Parole trovate nella pagina'], start=1):
    c = ws.cell(row=1, column=j, value=h); c.font = H_FONT; c.fill = H_FILL
for i,r in enumerate(idx, start=2):
    for j,v in enumerate([r['insegna'], r['validita'], PDF.get(r['chiave'], r['chiave']),
                          r['pagina'], r['parole']], start=1):
        c = ws.cell(row=i, column=j, value=v); c.font = BODY
        c.alignment = Alignment(vertical='top', wrap_text=(j==5))
for col,w in zip('ABCDE',[16,42,44,8,110]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:E{len(idx)+1}'

wb.save('out/offerte-supermercati-torino.xlsx')
print('prodotti:', len(PRODOTTI), '| pagine indicizzate:', len(idx))
