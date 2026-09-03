"""Inserisce nel file .xlsx il valore gia calcolato accanto a ogni formula.
Serve perche LibreOffice qui non parte e openpyxl non scrive i valori in cache:
senza, alcuni lettori (soprattutto su telefono) mostrerebbero la colonna vuota."""
import zipfile, shutil, re, os
from openpyxl import load_workbook

SRC = 'out/offerte-supermercati-torino.xlsx'
wb = load_workbook(SRC)
ws = wb['I tuoi prodotti']

# calcola in Python quello che la formula calcolerebbe: prezzo / quantita
attesi = {}
for r in range(2, ws.max_row + 1):
    q, p = ws.cell(r, 7).value, ws.cell(r, 8).value
    if isinstance(q, (int, float)) and isinstance(p, (int, float)) and q:
        attesi[f'I{r}'] = round(p / q, 6)
print('formule da valorizzare:', len(attesi))

# trova il file XML del foglio giusto
idx = wb.sheetnames.index('I tuoi prodotti') + 1
target = f'xl/worksheets/sheet{idx}.xml'

tmp = SRC + '.tmp'
with zipfile.ZipFile(SRC) as zin, zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
    for it in zin.infolist():
        data = zin.read(it.filename)
        if it.filename == target:
            x = data.decode('utf-8')
            def sost(m):
                ref, dentro = m.group(1), m.group(0)
                if ref in attesi and '<f>' in dentro and '<v>' not in dentro:
                    return dentro.replace('</f>', f'</f><v>{attesi[ref]}</v>')
                return dentro
            x, n = re.subn(r'<c r="([A-Z]+\d+)"[^>]*>.*?</c>', sost, x, flags=re.S)
            data = x.encode('utf-8')
        zout.writestr(it, data)
shutil.move(tmp, SRC)

# verifica: rileggo i valori in cache
wb2 = load_workbook(SRC, data_only=True)
ws2 = wb2['I tuoi prodotti']
vuoti = [r for r in range(2, ws2.max_row + 1) if ws2.cell(r, 9).value is None]
print('celle EUR/kg ancora vuote:', vuoti if vuoti else 'nessuna')
for r in (2, 3, 12, 20):
    print(f'  riga {r}: {ws2.cell(r,5).value[:38]:38s} -> {ws2.cell(r,9).value} EUR/kg')
